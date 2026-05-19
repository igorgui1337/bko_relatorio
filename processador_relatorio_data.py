#!/usr/bin/env python3
"""
processador_relatorio_data.py

Processador de dados de tickets BKO.

Entrada : CSV limpo gerado pelo validador_tabela_ticket.py
Saída   : XLSX com múltiplas abas de análise

Pipeline: validador_tabela_ticket.py → processador_relatorio_data.py → Dashboard (Streamlit)

Abas geradas:
  Aba_Geral         — Um registro por ticket com todos os indicadores
  SLA_Em_Processo   — Tickets abertos / em processo ordenados por tempo em aberto
  SLA_Resposta      — Tempo de resposta BO → BP (analista → consultor) por ticket
  Funil_Mensal      — Abertos / Em processo / Fechados por mês
  Funil_Semanal     — Abertos / Em processo / Fechados por semana
  Por_Assunto       — Agrupamento por assunto do ticket
  Por_Escritorio    — [OCULTO - sem dept] Agrupamento por escritório

Uso:
    python processador_relatorio_data.py <arquivo_ETL.csv>
    python processador_relatorio_data.py <arquivo_ETL.csv> <relatorio.xlsx>
"""

import re
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Caracteres ilegais para células Excel (controles ASCII exceto tab/LF/CR)
_ILLEGAL_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")

# ---------------------------------------------------------------------------
# Configurações
# ---------------------------------------------------------------------------

AGORA = datetime.now()

# Pares (coluna_data, coluna_hora) → coluna datetime resultante
DATETIME_PAIRS = [
    ("open_at",             "hora_open",    "dt_abertura"),
    ("answered_at",         "answered_hora","dt_resposta"),
    ("message_at",          "message_hora", "dt_mensagem"),
    ("previous_message_at", "previous_hora","dt_anterior"),
]

# Abas ocultas (nenhuma no momento — departamentos serão adicionados depois)
HIDDEN_SHEETS: set[str] = set()

# Estilos XLSX
HEADER_COLOR  = "1F3864"   # azul escuro
OPEN_COLOR    = "FFF2CC"   # amarelo claro → ticket aberto
PROC_COLOR    = "DEEBF7"   # azul claro → processing
CLOSED_COLOR  = "E2EFDA"   # verde claro → closed
ALERT_COLOR   = "FCE4D6"   # laranja → SLA estourado (> 24 h)

SLA_ALERTA_H = 24           # horas → acima disso é alerta no SLA


# ---------------------------------------------------------------------------
# 1. Carregamento
# ---------------------------------------------------------------------------

def _parse_datetimes(df: pd.DataFrame) -> pd.DataFrame:
    for date_col, hora_col, new_col in DATETIME_PAIRS:
        if date_col in df.columns and hora_col in df.columns:
            combined = df[date_col].str.strip() + " " + df[hora_col].str.strip()
            df[new_col] = pd.to_datetime(combined, format="%d/%m/%Y %H:%M", errors="coerce")
    return df


def load_csv(path: str) -> pd.DataFrame:
    """Carrega o CSV do ETL (UTF-8-BOM ou UTF-8) e cria colunas datetime."""
    encodings = ["utf-8-sig", "utf-8", "latin-1"]
    df = None
    for enc in encodings:
        try:
            df = pd.read_csv(path, sep=";", encoding=enc, dtype=str, keep_default_na=False)
            break
        except UnicodeDecodeError:
            continue
    if df is None:
        raise ValueError(f"Não foi possível ler o arquivo: {path}")

    str_cols = df.select_dtypes(include=["object", "str"]).columns
    df[str_cols] = df[str_cols].apply(lambda s: s.str.strip())
    df = _parse_datetimes(df)
    return df


# ---------------------------------------------------------------------------
# 2. Consolidação — um registro por ticket
# ---------------------------------------------------------------------------

def _first_nonempty(s: pd.Series) -> str:
    vals = s[s.ne("")]
    return vals.iloc[0] if len(vals) else ""


def _last_nonempty(s: pd.Series) -> str:
    vals = s[s.ne("")]
    return vals.iloc[-1] if len(vals) else ""


def _resolve_status(s: pd.Series) -> str:
    """closed > processing > open — o status mais crítico vence."""
    vals = set(s.str.lower().unique())
    for priority in ("closed", "processing", "open"):
        if priority in vals:
            return priority
    return s.iloc[-1]


def _count_transfers(group: pd.DataFrame) -> int:
    """
    Transferência = linha onde previous_sender_id ≠ "" e ≠ sender_id.
    Representa troca real de responsável dentro do ticket.
    """
    mask = (
        group["previous_sender_id"].ne("") &
        group["previous_sender_id"].ne(group["sender_id"])
    )
    return int(mask.sum())


def consolidate(df: pd.DataFrame) -> pd.DataFrame:
    """
    Agrega N linhas por ticket em 1 registro consolidado.

    Lógica das datas (open_at e answered_at são constantes por ticket):
      - dt_abertura   : min — garante a primeira abertura
      - dt_resposta   : min não-nula — primeira resposta (SLA BO > BP)
      - dt_ultima_msg : max(dt_mensagem) — última atividade registrada

    SLA em processo:
      - open / processing : AGORA - dt_abertura
      - closed            : dt_ultima_msg - dt_abertura

    SLA resposta BO > BP:
      - dt_resposta - dt_abertura (horas)
    """
    records = []

    for ticket_id, grp in df.groupby("ticket_id", sort=False):
        grp = grp.sort_values("dt_mensagem", na_position="last")

        dt_abertura   = grp["dt_abertura"].min()
        dt_resposta   = grp["dt_resposta"].dropna().min() if "dt_resposta" in grp else pd.NaT
        dt_ultima_msg = grp["dt_mensagem"].dropna().max()

        # Fallback: se não há dt_mensagem, usa dt_resposta como última atividade
        if pd.isna(dt_ultima_msg):
            dt_ultima_msg = dt_resposta

        status = _resolve_status(grp["status"])

        # Tempo em processo (horas)
        if status in ("open", "processing"):
            ref_fim = AGORA
        else:
            ref_fim = dt_ultima_msg if pd.notna(dt_ultima_msg) else dt_abertura

        tempo_processo_h = (
            round((ref_fim - dt_abertura).total_seconds() / 3600, 2)
            if pd.notna(dt_abertura) else None
        )

        # SLA resposta BO > BP (horas)
        tempo_resposta_h = (
            round((dt_resposta - dt_abertura).total_seconds() / 3600, 2)
            if (pd.notna(dt_resposta) and pd.notna(dt_abertura)) else None
        )

        records.append({
            "ticket_id"         : ticket_id,
            "ticket_subject"    : _first_nonempty(grp["ticket_subject"]),
            "status"            : status,
            "dt_abertura"       : dt_abertura,
            "dt_resposta"       : dt_resposta,
            "dt_ultima_atividade": dt_ultima_msg,
            "tempo_processo_h"  : tempo_processo_h,
            "tempo_resposta_h"  : tempo_resposta_h,
            "sla_alerta"        : "SIM" if (tempo_processo_h or 0) > SLA_ALERTA_H and status != "closed" else "",
            "n_mensagens"       : len(grp),
            "n_transferencias"  : _count_transfers(grp),
            "analista"          : _last_nonempty(grp["sender"]),
            "analista_id"       : _last_nonempty(grp["sender_id"]),
            "consultor"         : _last_nonempty(grp["consultant"]),
            "consultor_id"      : _last_nonempty(grp["consultant_id"]),
            "escritorio"        : _last_nonempty(grp["office"]),
            "mes_abertura"      : dt_abertura.strftime("%Y-%m") if pd.notna(dt_abertura) else "",
            "semana_abertura"   : (
                f"{dt_abertura.year}-S{dt_abertura.isocalendar()[1]:02d}"
                if pd.notna(dt_abertura) else ""
            ),
        })

    return pd.DataFrame(records)


# ---------------------------------------------------------------------------
# 3. Construção das abas analíticas
# ---------------------------------------------------------------------------

def _agg_funil(df: pd.DataFrame, group_col: str) -> pd.DataFrame:
    grouped = df.groupby(group_col).agg(
        tickets_abertos          =("ticket_id",      "count"),
        tickets_fechados         =("status",          lambda s: (s == "closed").sum()),
        em_processo              =("status",          lambda s: s.isin(["open", "processing"]).sum()),
        com_sla_alerta           =("sla_alerta",      lambda s: (s == "SIM").sum()),
        tempo_medio_processo_h   =("tempo_processo_h","mean"),
        tempo_medio_resposta_h   =("tempo_resposta_h","mean"),
        n_transferencias_total   =("n_transferencias","sum"),
    ).reset_index()

    grouped.rename(columns={group_col: "periodo"}, inplace=True)
    grouped["pct_fechado"]              = (grouped["tickets_fechados"] / grouped["tickets_abertos"] * 100).round(1)
    grouped["tempo_medio_processo_h"]   = grouped["tempo_medio_processo_h"].round(2)
    grouped["tempo_medio_resposta_h"]   = grouped["tempo_medio_resposta_h"].round(2)
    return grouped.sort_values("periodo")


def make_sla_em_processo(df: pd.DataFrame) -> pd.DataFrame:
    mask = df["status"].isin(["open", "processing"])
    cols = [
        "ticket_id", "ticket_subject", "status", "sla_alerta",
        "dt_abertura", "tempo_processo_h",
        "n_mensagens", "n_transferencias",
        "analista", "consultor", "escritorio",
    ]
    return (
        df.loc[mask, cols]
        .sort_values("tempo_processo_h", ascending=False)
        .reset_index(drop=True)
    )


def make_sla_resposta(df: pd.DataFrame) -> pd.DataFrame:
    mask = df["tempo_resposta_h"].notna()
    cols = [
        "ticket_id", "ticket_subject", "status",
        "dt_abertura", "dt_resposta", "tempo_resposta_h",
        "analista", "consultor", "escritorio",
    ]
    return (
        df.loc[mask, cols]
        .sort_values("tempo_resposta_h", ascending=False)
        .reset_index(drop=True)
    )


def make_funil_mensal(df: pd.DataFrame) -> pd.DataFrame:
    return _agg_funil(df, "mes_abertura")


def make_funil_semanal(df: pd.DataFrame) -> pd.DataFrame:
    return _agg_funil(df, "semana_abertura")


def make_por_assunto(df: pd.DataFrame) -> pd.DataFrame:
    grouped = df.groupby("ticket_subject").agg(
        total                    =("ticket_id",       "count"),
        fechados                 =("status",           lambda s: (s == "closed").sum()),
        em_processo              =("status",           lambda s: s.isin(["open", "processing"]).sum()),
        com_sla_alerta           =("sla_alerta",       lambda s: (s == "SIM").sum()),
        tempo_medio_resposta_h   =("tempo_resposta_h", "mean"),
        tempo_medio_processo_h   =("tempo_processo_h", "mean"),
        n_transferencias_medio   =("n_transferencias", "mean"),
    ).reset_index()
    grouped.rename(columns={"ticket_subject": "assunto"}, inplace=True)
    grouped["pct_fechado"]            = (grouped["fechados"] / grouped["total"] * 100).round(1)
    grouped["tempo_medio_resposta_h"] = grouped["tempo_medio_resposta_h"].round(2)
    grouped["tempo_medio_processo_h"] = grouped["tempo_medio_processo_h"].round(2)
    grouped["n_transferencias_medio"] = grouped["n_transferencias_medio"].round(1)
    return grouped.sort_values("total", ascending=False).reset_index(drop=True)


def make_por_escritorio(df: pd.DataFrame) -> pd.DataFrame:
    grouped = df.groupby("escritorio").agg(
        total_tickets            =("ticket_id",       "count"),
        fechados                 =("status",           lambda s: (s == "closed").sum()),
        pendentes                =("status",           lambda s: s.isin(["open", "processing", "pending"]).sum()),
        com_sla_alerta           =("sla_alerta",       lambda s: (s == "SIM").sum()),
        tempo_medio_resposta_h   =("tempo_resposta_h", "mean"),
        tempo_min_resposta_h     =("tempo_resposta_h", "min"),
        tempo_max_resposta_h     =("tempo_resposta_h", "max"),
        tempo_medio_processo_h   =("tempo_processo_h", "mean"),
    ).reset_index()
    grouped["pct_fechado"]            = (grouped["fechados"]  / grouped["total_tickets"] * 100).round(1)
    grouped["pct_pendente"]           = (grouped["pendentes"] / grouped["total_tickets"] * 100).round(1)
    grouped["tempo_medio_resposta_h"] = grouped["tempo_medio_resposta_h"].round(2)
    grouped["tempo_min_resposta_h"]   = grouped["tempo_min_resposta_h"].round(2)
    grouped["tempo_max_resposta_h"]   = grouped["tempo_max_resposta_h"].round(2)
    grouped["tempo_medio_processo_h"] = grouped["tempo_medio_processo_h"].round(2)

    # Ordem de colunas para leitura natural
    col_order = [
        "escritorio",
        "total_tickets",
        "fechados",        "pct_fechado",
        "pendentes",       "pct_pendente",
        "com_sla_alerta",
        "tempo_medio_resposta_h",
        "tempo_min_resposta_h",
        "tempo_max_resposta_h",
        "tempo_medio_processo_h",
    ]
    return grouped[col_order].sort_values("total_tickets", ascending=False).reset_index(drop=True)


# ---------------------------------------------------------------------------
# 4. Escrita do XLSX com formatação
# ---------------------------------------------------------------------------

_STATUS_FILL = {
    "open":       PatternFill(start_color=OPEN_COLOR,   end_color=OPEN_COLOR,   fill_type="solid"),
    "processing": PatternFill(start_color=PROC_COLOR,   end_color=PROC_COLOR,   fill_type="solid"),
    "closed":     PatternFill(start_color=CLOSED_COLOR, end_color=CLOSED_COLOR, fill_type="solid"),
}
_ALERT_FILL = PatternFill(start_color=ALERT_COLOR, end_color=ALERT_COLOR, fill_type="solid")


def _apply_header(ws) -> None:
    fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20


def _autowidth(ws, df: pd.DataFrame) -> None:
    for col_idx, col in enumerate(df.columns, 1):
        letter = get_column_letter(col_idx)
        max_len = max(
            len(str(col)),
            df[col].astype(str).str.len().max() if len(df) else 0,
        )
        ws.column_dimensions[letter].width = min(max_len + 4, 55)


_COLOR_LIMIT = 500  # aplica coloracao apenas em sheets com ate N linhas


def _color_status_rows(ws, df: pd.DataFrame, status_col: str) -> None:
    """
    Pinta cada linha conforme o status do ticket.
    Ignorado para sheets grandes (> _COLOR_LIMIT linhas) por performance.
    """
    if status_col not in df.columns or len(df) > _COLOR_LIMIT:
        return

    alert_col_idx = df.columns.get_loc("sla_alerta") + 1 if "sla_alerta" in df.columns else None

    status_vals = df[status_col].str.lower().tolist()
    alert_vals  = (df["sla_alerta"] == "SIM").tolist() if "sla_alerta" in df.columns else [False] * len(df)

    for i, (status_val, is_alert) in enumerate(zip(status_vals, alert_vals), start=2):
        fill       = _STATUS_FILL.get(status_val)
        alert_fill = _ALERT_FILL if is_alert else None
        if not fill and not alert_fill:
            continue
        for cell in ws[i]:
            if alert_fill and cell.column == alert_col_idx:
                cell.fill = alert_fill
            elif fill:
                cell.fill = fill


def _sanitize_df(df: pd.DataFrame) -> pd.DataFrame:
    """Converte datetime para string HH:MM (sem segundos) e remove chars ilegais."""
    dt_cols = df.select_dtypes(include=["datetime64[ns]", "datetimetz"]).columns
    for col in dt_cols:
        df[col] = df[col].dt.strftime("%d/%m/%Y %H:%M").fillna("")

    for col in df.select_dtypes(include=["object", "str"]).columns:
        df[col] = df[col].astype(str).apply(
            lambda v: _ILLEGAL_CHARS.sub("", v) if isinstance(v, str) else v
        )
    return df


# ---------------------------------------------------------------------------
# 4b. Gráficos no XLSX
# ---------------------------------------------------------------------------

# Cores das series (hex sem #) — mesmas do dashboard Streamlit
_C_FECHADO   = "2CA02C"
_C_PROCESSO  = "FF7F0E"
_C_ABERTO    = "1F77B4"
_C_ALERTA    = "D62728"
_C_PRINCIPAL = "1F3864"


def _set_color(series, fill_hex: str) -> None:
    try:
        series.graphicalProperties.solidFill = fill_hex
        series.graphicalProperties.line.solidFill = fill_hex
    except Exception:
        pass


def _add_labels(chart, show_pct: bool = False) -> None:
    """Adiciona rótulos de dados em um gráfico openpyxl."""
    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = True
    chart.dLbls.showCatName = False
    chart.dLbls.showSerName = False
    chart.dLbls.showLegendKey = False
    chart.dLbls.showPercent = show_pct


def _col(df: pd.DataFrame, name: str) -> int:
    return df.columns.get_loc(name) + 1


def _chart_funil(ws, df: pd.DataFrame) -> None:
    """Barras agrupadas (abertos/fechados/em_processo) + linha de SLA por periodo."""
    n = len(df) + 1
    cats = Reference(ws, min_col=_col(df, "periodo"), min_row=2, max_row=n)
    chart_row = n + 3

    # Barras horizontais — apenas fechados e em_processo
    bar = BarChart()
    bar.type = "bar"
    bar.grouping = "clustered"
    bar.title = "Funil de Tickets"
    bar.x_axis.title = "Quantidade"
    bar.width = 22
    bar.height = 14

    for col_name, color in [
        ("tickets_fechados", _C_FECHADO),
        ("em_processo",      _C_PROCESSO),
    ]:
        ref = Reference(ws, min_col=_col(df, col_name), max_col=_col(df, col_name),
                        min_row=1, max_row=n)
        bar.add_data(ref, titles_from_data=True)
        _set_color(bar.series[-1], color)

    bar.set_categories(cats)
    _add_labels(bar)
    ws.add_chart(bar, f"A{chart_row}")

    # Line: tempo medio resposta
    line = LineChart()
    line.title = "Tempo Medio de Resposta (h)"
    line.y_axis.title = "Horas"
    line.width = 18
    line.height = 12

    ref_resp = Reference(ws, min_col=_col(df, "tempo_medio_resposta_h"),
                         max_col=_col(df, "tempo_medio_resposta_h"),
                         min_row=1, max_row=n)
    line.add_data(ref_resp, titles_from_data=True)
    line.set_categories(cats)
    _add_labels(line)
    try:
        line.series[0].graphicalProperties.line.solidFill = _C_PRINCIPAL
        line.series[0].graphicalProperties.line.width = 25000
        line.series[0].marker.symbol = "circle"
        line.series[0].marker.size = 6
    except Exception:
        pass

    ws.add_chart(line, f"{get_column_letter(len(df.columns) + 2)}{chart_row}")


def _chart_por_assunto(ws, df: pd.DataFrame) -> None:
    """Bar horizontal: top 20 assuntos por volume com % fechado."""
    n_plot = min(20, len(df))
    n = n_plot + 1
    chart_row = len(df) + 4

    bar = BarChart()
    bar.type = "bar"
    bar.grouping = "clustered"
    bar.title = "Top 20 Assuntos por Volume"
    bar.x_axis.title = "Quantidade"
    bar.width = 26
    bar.height = 16

    for col_name, color in [
        ("total",    _C_PRINCIPAL),
        ("fechados", _C_FECHADO),
        ("em_processo", _C_PROCESSO),
    ]:
        ref = Reference(ws, min_col=_col(df, col_name), max_col=_col(df, col_name),
                        min_row=1, max_row=n)
        bar.add_data(ref, titles_from_data=True)
        _set_color(bar.series[-1], color)

    cats = Reference(ws, min_col=_col(df, "assunto"), min_row=2, max_row=n)
    bar.set_categories(cats)
    _add_labels(bar)
    ws.add_chart(bar, f"A{chart_row}")


def _chart_por_escritorio(ws, df: pd.DataFrame) -> None:
    """Barras: fechados vs pendentes + barra de tempo medio de resposta."""
    n_plot = min(20, len(df))
    n = n_plot + 1
    chart_row = len(df) + 4

    cats = Reference(ws, min_col=_col(df, "escritorio"), min_row=2, max_row=n)

    # Fechados vs pendentes
    bar = BarChart()
    bar.type = "bar"
    bar.grouping = "clustered"
    bar.title = "Fechados vs Pendentes — Top 20 Escritorios"
    bar.x_axis.title = "Quantidade"
    bar.width = 24
    bar.height = 16

    for col_name, color in [
        ("fechados",  _C_FECHADO),
        ("pendentes", _C_ALERTA),
    ]:
        ref = Reference(ws, min_col=_col(df, col_name), max_col=_col(df, col_name),
                        min_row=1, max_row=n)
        bar.add_data(ref, titles_from_data=True)
        _set_color(bar.series[-1], color)

    bar.set_categories(cats)
    _add_labels(bar)
    ws.add_chart(bar, f"A{chart_row}")

    # Tempo medio resposta
    bar2 = BarChart()
    bar2.type = "bar"
    bar2.title = "Tempo Medio de Resposta (h) — Top 20 Escritorios"
    bar2.x_axis.title = "Horas"
    bar2.width = 22
    bar2.height = 14

    ref_resp = Reference(ws, min_col=_col(df, "tempo_medio_resposta_h"),
                         max_col=_col(df, "tempo_medio_resposta_h"),
                         min_row=1, max_row=n)
    bar2.add_data(ref_resp, titles_from_data=True)
    bar2.set_categories(cats)
    _set_color(bar2.series[0], _C_PROCESSO)
    _add_labels(bar2)

    ws.add_chart(bar2, f"{get_column_letter(len(df.columns) + 2)}{chart_row}")


def _chart_aba_geral(ws, df: pd.DataFrame) -> None:
    """Pie chart de status a partir de mini tabela adicionada abaixo dos dados."""
    n_data = len(df) + 1
    summary_row = n_data + 3

    # Escreve mini tabela de contagem
    status_counts = df["status"].value_counts()
    ws.cell(row=summary_row, column=1, value="Status")
    ws.cell(row=summary_row, column=2, value="Total")
    for i, (status, count) in enumerate(status_counts.items(), start=1):
        ws.cell(row=summary_row + i, column=1, value=str(status))
        ws.cell(row=summary_row + i, column=2, value=int(count))

    n_s = len(status_counts)

    # Pie chart
    pie = PieChart()
    pie.title = "Distribuicao por Status"
    pie.width = 16
    pie.height = 14

    labels = Reference(ws, min_col=1, min_row=summary_row + 1, max_row=summary_row + n_s)
    data   = Reference(ws, min_col=2, min_row=summary_row,     max_row=summary_row + n_s)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    _add_labels(pie, show_pct=True)

    colors_status = [_C_FECHADO, _C_PROCESSO, _C_ABERTO, _C_ALERTA]
    for i, color in enumerate(colors_status[:n_s]):
        pt = DataPoint(idx=i)
        try:
            pt.graphicalProperties.solidFill = color
        except Exception:
            pass
        pie.series[0].dPt.append(pt)

    ws.add_chart(pie, f"D{summary_row}")


def _add_charts_to_workbook(writer, sheets: dict[str, pd.DataFrame]) -> None:
    """Adiciona gráficos em cada aba após a escrita dos dados."""
    handlers = {
        "Funil_Mensal":    _chart_funil,
        "Funil_Semanal":   _chart_funil,
        "Por_Assunto":     _chart_por_assunto,
        "Por_Escritorio":  _chart_por_escritorio,
        "Aba_Geral":       _chart_aba_geral,
    }
    for sheet_name, fn in handlers.items():
        if sheet_name not in sheets or len(sheets[sheet_name]) == 0:
            continue
        try:
            fn(writer.sheets[sheet_name], sheets[sheet_name])
        except Exception as e:
            print(f"  [aviso] grafico '{sheet_name}' nao gerado: {e}")


def write_xlsx(path: str, sheets: dict[str, pd.DataFrame]) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df = _sanitize_df(df.copy())
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            _apply_header(ws)
            _autowidth(ws, df)
            ws.freeze_panes = "A2"
            if "status" in df.columns:
                _color_status_rows(ws, df, "status")
            if sheet_name in HIDDEN_SHEETS:
                ws.sheet_state = "hidden"

        _add_charts_to_workbook(writer, sheets)

    print(f"  XLSX salvo: {Path(path).name}")


# ---------------------------------------------------------------------------
# 5. Pipeline principal
# ---------------------------------------------------------------------------

def process(input_path: str, output_path: str | None = None) -> str:
    p_in = Path(input_path)
    if not p_in.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {p_in}")

    if output_path is None:
        stem = p_in.stem.replace("_ETL", "")
        output_path = p_in.parent / f"{stem}_relatorio.xlsx"
    p_out = Path(output_path)

    print(f"\n{'='*60}")
    print(f"  PROCESSADOR DE RELATÓRIO BKO")
    print(f"  Referência: {AGORA.strftime('%d/%m/%Y %H:%M')}")
    print(f"{'='*60}")
    print(f"  Entrada : {p_in.name}")

    # 1. Carregar
    df_raw = load_csv(str(p_in))
    print(f"  Linhas  : {len(df_raw):,}")

    # 2. Consolidar por ticket
    df_geral = consolidate(df_raw)
    n_tickets = len(df_geral)
    status_cnt = df_geral["status"].value_counts().to_dict()
    print(f"  Tickets : {n_tickets:,}  |  {status_cnt}")
    print(f"  SLA alerta (>{SLA_ALERTA_H}h): {(df_geral['sla_alerta'] == 'SIM').sum()}")

    # 3. Construir abas
    df_sla_proc = make_sla_em_processo(df_geral)
    df_sla_resp = make_sla_resposta(df_geral)
    df_funil_m  = make_funil_mensal(df_geral)
    df_funil_s  = make_funil_semanal(df_geral)
    df_assunto  = make_por_assunto(df_geral)
    df_escrit   = make_por_escritorio(df_geral)

    # 4. Escrever XLSX
    sheets = {
        "Aba_Geral"        : df_geral,
        "SLA_Em_Processo"  : df_sla_proc,
        "SLA_Resposta"     : df_sla_resp,
        "Funil_Mensal"     : df_funil_m,
        "Funil_Semanal"    : df_funil_s,
        "Por_Assunto"      : df_assunto,
        "Por_Escritorio"   : df_escrit,    # hidden até receber departamentos
    }
    write_xlsx(str(p_out), sheets)

    print()
    print(f"  {'Aba':<20} {'Registros':>10}")
    print(f"  {'-'*32}")
    for name, df in sheets.items():
        hidden = " (oculta)" if name in HIDDEN_SHEETS else ""
        print(f"  {name:<20} {len(df):>10,}{hidden}")
    print(f"{'='*60}\n")

    return str(p_out)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        print("Uso: python processador_relatorio_data.py <arquivo_ETL.csv> [relatorio.xlsx]")
        sys.exit(1)

    inp = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else None

    try:
        process(inp, out)
    except (FileNotFoundError, ValueError) as e:
        print(f"ERRO: {e}")
        sys.exit(1)
